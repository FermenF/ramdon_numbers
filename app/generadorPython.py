import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from colorama import Fore
from app.constants import path_open_route as path, open_file

book = Workbook();
sheet = book.active;
data = [];
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

n = 50000

def generate_numbers():
    c = 0
    for i in range(n):
        i+=1
        c+=1
        numero = np.random.rand()

        if(c > 10000):
            c = 1
        if(i <= 10000):
            excel_location = ("A"+str(c))
        if(i > 10000 and i <= 20000):
            excel_location = ("B"+str(c))
        if(i > 20000 and i <= 30000):
            excel_location = ("C"+str(c))
        if(i > 30000 and i <= 40000):
            excel_location = ("D"+str(c))
        if(i > 40000 and i <= 50000):
            excel_location = ("E"+str(c))

        sheet[excel_location] = numero
        data.append(numero)

    # N 
    sheet["H2"] = 'N'
    sheet["I2"] = n

    # K
    sheet["H3"] = 'K'
    sheet["I3"] = 'REDONDEAR(1,33*LN(I2)+0,5;0)'
    sheet["I3"].fill = redFill

    # Rango
    sheet["H4"] = 'Rango'
    sheet["I4"] = '1'

    # Amplitud
    sheet["H5"] = 'Amplitud'
    sheet["I5"] = '=I4/I3'

    # Chi-cuadrado-crit
    sheet["H6"] = 'Chi-cuadrado-crit'
    sheet["I6"] = 'DISTR.CHICUAD.CD(1-I7;I3-1)'
    sheet["I6"].fill = redFill
    
    # Alpha
    sheet["H7"] = 'Alpha'
    sheet["I7"] = '5%'

    sheet["L2"] = 'Una vez generado los cambios, guarde y cierre el archivo, ejecute el programa nuevamente y seleccione la opción 2.'

    print(Fore.GREEN + "\nEjecución de programa..." + Fore.RESET)

    try:
        book.save('./app/docs/num_random.xlsx')
        print(Fore.LIGHTBLUE_EX + "\nDocumento generado correctamente, debe realizar modificaciones al documento para continuar..." +  Fore.RESET)
        print(Fore.YELLOW + '*** Agregue el simbolo "=" antes del texto, en las casillas de color '+ Fore.RESET + Fore.RED + 'rojo' + Fore.RESET)
        
        open_file(path)
        
    except:
        print(Fore.RED + "\nError al generar documento, cierre el archivo he intente nuevamente." + Fore.RESET)