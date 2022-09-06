import openpyxl
from app.constants import path_open_route as path, open_file
from openpyxl.styles import PatternFill
from app.generadorPython import data as data_numbers
import matplotlib.pyplot as plt

book = openpyxl.load_workbook(path, data_only=True)
sheet = book.active
normalFill = PatternFill(fill_type=None, start_color='FFFFFFFF', end_color='FF000000')
redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
greenFill = PatternFill(start_color='0099CC00', end_color='0099CC00', fill_type='solid')	

k = sheet.cell(row = 3, column = 9)

def checkDataChange():
    if type(k.value) == int or type(k.value) == float:
        return generateTable()
    return print("No se han detectado cambios, guarde el archivo o intente de nuevo")

def generateTable():

    sheet['L2'] = ''
    sheet["I3"].fill = normalFill
    sheet["I6"].fill = normalFill

    print("\nGenerando tabla de frecuencia...")

    # Cabecera
    sheet['H10'] = 'Li'
    sheet['I10'] = 'Ls'
    sheet['J10'] = 'Fe'
    sheet['K10'] = 'Fo'
    sheet['L10'] = 'ECM'

    # Data inicial
    sheet['H11'] = '0'
    sheet['I11'] = '=I5'
    sheet['J11'] = '=$I$2/$I$3'

    ls = sheet.cell(row = 5, column = 9)
    sheet['K11'] = 'CONTAR.SI($A:$E;"<"&I11)'
    sheet['K11'].fill = redFill

    sheet['L11'] = '=((J11-K11)/J11)^2'

    print("\nCabecera de la tabla generada.")

    row = 12
    row_ls = 11
    for i in range(k.value-1):

        row_location = ("H"+str(row))
        row_loc_ls = ("I"+str(row_ls))
        column_ls = ("I"+str(row))
        column_fe = ("J"+str(row))

        sheet[column_fe] = '=$I$2/$I$3'
        sheet[column_ls] = "=$I$5+"+row_location 
        sheet[row_location] = "="+row_loc_ls

        row+=1
        row_ls+=1
    
    sheet["K12"] = 'CONTAR.SI($A:$E;"<"&I12)-SUMA($K$11:K11)'
    sheet["K12"].fill = greenFill

    sheet["K2"] = 'Promedio'
    sheet["L2"] = '=Promedio(A1:E10000)'

    sheet["J27"] = 'Suma'
    sheet["K27"] = '=SUMA(K11:K25)'

    sheet["N10"] = 'Agregue el signo = antes de las formulas (COLOR ROJO)'
    sheet["N11"] = 'Agregue el signo = antes de las formulas COLOR VERDE Y ARRASTRE HACIA ABAJO - SEGÚN TABLA'

    book.save(path)

    print("\nDatos generados.")

    open_file(path)

def histogram(data, n_bins):
    option = str(input("Desea ver un histogrma de graficos? S/N: "))
    if(option == "S" or option == "s"):
        plt.hist(data, bins=n_bins, density=True, color="yellow")
        plt.xlabel("Values")
        plt.ylabel("Frequency")
        plt.title("Histogram")
        plt.show()